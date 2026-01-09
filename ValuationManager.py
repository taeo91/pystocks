import logging
import datetime
import os
import pandas as pd
from openpyxl.styles import PatternFill
from AppManager import get_db_connection

class ValuationManager:
    """
    주식 가치 평가를 관리하는 클래스
    """
    # 평가 결과 상수
    _RESULT_UNDERVALUED = "저평가"
    _RESULT_OVERVALUED = "고평가"
    _RESULT_FAIR = "적정"
    _RESULT_NA = "N/A"
 
    def __init__(self, db_access):
        """
        ValuationManager 생성자

        Args:
            db_access: 데이터베이스 접근 객체
        """
        self.db_access = db_access
        # RIM 모델 상수 설정
        self.REQUIRED_ROE = float(os.getenv('VAL_REQUIRED_ROE', '8.0'))
        self.LOW_VALUATION_THRESHOLD = float(os.getenv('VAL_LOW_THRESHOLD', '-10.0'))
        self.HIGH_VALUATION_THRESHOLD = float(os.getenv('VAL_HIGH_THRESHOLD', '10.0'))
        self.CONSERVATIVE_FACTOR = float(os.getenv('VAL_CONSERVATIVE_FACTOR', '0.8'))
        
        # 가치평가 모델 가중치
        self.W_RIM = float(os.getenv('VAL_W_RIM', '0.6'))
        self.W_PER = float(os.getenv('VAL_W_PER', '0.2'))
        self.W_PEGR = float(os.getenv('VAL_W_PEGR', '0.2'))

        # PEGR 모델 적용 조건
        self.PEGR_MIN_GROWTH = float(os.getenv('VAL_PEGR_MIN_GROWTH', '5'))
        self.PEGR_MAX_GROWTH = float(os.getenv('VAL_PEGR_MAX_GROWTH', '50'))

    def create_valuation_table(self):
        """'valuations' 테이블을 생성하는 메서드"""
        try:
            query = """
            CREATE TABLE IF NOT EXISTS valuations (
                id INT AUTO_INCREMENT PRIMARY KEY,
                code VARCHAR(20) NOT NULL COMMENT '종목코드',
                date DATE NOT NULL COMMENT '평가일자',
                fair_value DECIMAL(15, 2) COMMENT '적정주가',
                current_price DECIMAL(15, 2) COMMENT '현재주가',
                discrepancy_ratio DECIMAL(10, 4) COMMENT '괴리율 (%)',
                eps_growth_rate DECIMAL(10, 4) COMMENT 'EPS 성장률 (%)',
                peg_ratio DECIMAL(10, 2) COMMENT 'PEG 비율',
                valuation_result VARCHAR(50) COMMENT '평가결과 (저평가/고평가)',
                FOREIGN KEY (code) REFERENCES companies (code) ON DELETE CASCADE,
                UNIQUE KEY (code, date)
            ) COMMENT '일일 주식 가치 평가 정보';
            """
            self.db_access.execute_query(query)
            logging.info("Table 'valuations' created or already exists.")
            return True
        except Exception as e:
            logging.error(f"Error creating 'valuations' table: {e}")
            return False

    def calculate_and_save_valuations(self, limit=None):
        """DB에 저장된 모든 종목의 가치를 평가하고 결과를 DB와 엑셀 파일로 저장합니다."""
        try:
            logging.info("가치 평가를 시작합니다...")
            self.create_valuation_table()
            today = datetime.date.today()
            today_str = today.strftime('%Y-%m-%d')

            all_data = self._fetch_valuation_data_bulk(today_str, limit)
            if not all_data:
                logging.warning(f"'{today_str}' 날짜로 평가할 데이터가 없습니다.")
                return

            valuation_results = []
            for stock_data in all_data:
                # 필수 데이터 확인 (None 이면 계산 불가)
                required_keys = ['code', 'roe_pred', 'bps_pred', 'current_price']
                if not all(stock_data.get(key) is not None for key in required_keys):
                    logging.debug(f"'{stock_data.get('code')}'에 대한 충분한 데이터가 없어 평가를 건너뜁니다.")
                    continue

                result = self._perform_valuation_calculation(stock_data)
                if result:
                    valuation_results.append(result)

            if valuation_results:
                logging.info(f"총 {len(valuation_results)}개 종목의 가치 평가를 완료했습니다.")
                self._save_results_to_excel(valuation_results, today_str)
                self._save_results_to_db(valuation_results, today_str)
            else:
                logging.info("가치 평가를 수행할 종목이 없습니다.")

        except Exception as e:
            logging.error(f"가치 평가 계산 및 저장 중 오류 발생: {e}")

    def calculate_valuation(self, code, date_str):
        """
        개별 종목의 가치를 평가하는 메서드 (단일 종목용).
        여러 종목을 평가할 때는 성능이 최적화된 calculate_and_save_valuations() 사용을 권장합니다.
        """
        try:
            # 1. DB에서 필요한 재무 정보 및 현재가 가져오기
            # 참고: 이 메서드는 단일 종목에 대해 여러 쿼리를 실행하므로 대량 처리에는 비효율적입니다.
            query_financials = """
            SELECT pbr, per, indust_per, eps, roe, bps, eps_pred, roe_pred, bps_pred, perf_yoy, perf_vs_3m_ago
            FROM daily_financials 
            WHERE code = %s AND date = %s
            """
            financial_data_tuple = self.db_access.fetch_one(query_financials, (code, date_str))
            if not financial_data_tuple or any(d is None for d in financial_data_tuple[1:4]): # eps_pred, roe_pred, bps_pred
                logging.debug(f"'{code}'에 대한 충분한 재무 데이터가 없습니다.")
                return None

            fin_cols = ['pbr', 'per', 'indust_per', 'eps', 'roe', 'bps', 'eps_pred', 'roe_pred', 'bps_pred', 'perf_yoy', 'perf_vs_3m_ago']
            financial_data = dict(zip(fin_cols, financial_data_tuple))

            query_price = """
            SELECT p.close_price, c.name
            FROM prices p
            JOIN companies c ON p.company_id = c.id
            WHERE c.code = %s
            ORDER BY p.trade_date DESC LIMIT 1
            """
            price_data = self.db_access.fetch_one(query_price, (code,))
            if not price_data or price_data[0] is None:
                logging.debug(f"'{code}'에 대한 가격 정보가 없습니다.")
                return None

            # 2. 데이터 취합 및 중앙 계산 메서드 호출
            stock_data = {
                'code': code,
                'name': price_data[1],
                'current_price': price_data[0],
                **financial_data
            }

            return self._perform_valuation_calculation(stock_data)

        except Exception as e:
            logging.error(f"'{code}' 가치 평가 중 오류: {e}")
            return None

    def _fetch_valuation_data_bulk(self, date_str, limit=None):
        """평가에 필요한 모든 종목의 재무 및 가격 데이터를 한 번의 쿼리로 가져옵니다."""
        # 참고: 이 쿼리는 MySQL 8.0+ 또는 ROW_NUMBER()를 지원하는 DB에서 최적의 성능을 보입니다.
        # DB가 윈도우 함수를 지원하지 않는 경우, 서브쿼리 등으로 최신 가격을 가져오는 로직 수정이 필요합니다.
        query = """
            SELECT
                df.code, c.name, df.pbr, df.per, df.indust_per, df.eps, df.roe, df.bps, df.eps_pred, df.roe_pred, df.bps_pred,
                df.perf_yoy, df.perf_vs_3m_ago,
                p_latest.close_price AS current_price
            FROM
                daily_financials AS df
            JOIN
                companies AS c ON df.code = c.code
            JOIN
                (
                    SELECT
                        company_id,
                        close_price,
                        ROW_NUMBER() OVER(PARTITION BY company_id ORDER BY trade_date DESC) as rn
                    FROM prices
                ) AS p_latest ON c.id = p_latest.company_id AND p_latest.rn = 1
            WHERE
                df.date = %s
                AND c.code LIKE '%%0'
        """
        params = [date_str]
        if limit:
            query += " LIMIT %s"
            params.append(int(limit))

        # fetch_all이 파라미터를 지원하고, 결과가 튜플 리스트라고 가정합니다.
        rows = self.db_access.fetch_all(query, params)
        if not rows:
            return []

        columns = [
            'code', 'name', 'pbr', 'per', 'indust_per', 'eps', 'roe', 'bps', 'eps_pred', 'roe_pred', 'bps_pred',
            'perf_yoy', 'perf_vs_3m_ago', 'current_price'
        ]
        return [dict(zip(columns, row)) for row in rows]

    def _prepare_data(self, stock_data):
        """데이터를 추출하고 숫자형으로 변환합니다."""
        data = {key: stock_data.get(key) for key in [
            'code', 'name', 'current_price', 'pbr', 'per', 'indust_per',
            'eps', 'roe', 'bps', 'eps_pred', 'roe_pred', 'bps_pred',
            'perf_yoy', 'perf_vs_3m_ago'
        ]}
        
        for key, value in data.items():
            if key in ['code', 'name']:
                continue
            if isinstance(value, (int, float)):
                continue
            try:
                # Decimal 타입을 float으로 변환
                data[key] = float(value)
            except (ValueError, TypeError, AttributeError):
                # 변환 실패 시 None으로 처리하지 않고 원래 값을 유지 (perf* 필드 등)
                pass
        return data

    def _calculate_growth_rates(self, data):
        """EPS, BPS, ROE 성장률을 계산합니다."""
        rates = {'eps_growth_rate': 0, 'bps_growth_rate': 0, 'roe_growth_rate': 0}
        
        if data.get('eps') and data['eps'] > 0 and data.get('eps_pred'):
            rates['eps_growth_rate'] = ((data['eps_pred'] - data['eps']) / data['eps']) * 100
        
        if data.get('bps') and data['bps'] > 0 and data.get('bps_pred'):
            rates['bps_growth_rate'] = ((data['bps_pred'] - data['bps']) / data['bps']) * 100
            
        if data.get('roe') and data['roe'] > 0 and data.get('roe_pred'):
            rates['roe_growth_rate'] = ((data['roe_pred'] - data['roe']) / data['roe']) * 100
            
        return rates

    def _calculate_fair_value_rim(self, data):
        """RIM 모델 기반 적정주가를 계산합니다."""
        roe_pred = data.get('roe_pred')
        bps_pred = data.get('bps_pred')
        if roe_pred is not None and bps_pred is not None and roe_pred >= 0 and bps_pred > 0:
            excess_profit = (roe_pred / 100 - self.REQUIRED_ROE / 100) * bps_pred
            return bps_pred + (excess_profit / (self.REQUIRED_ROE / 100))
        return 0

    def _calculate_fair_value_per(self, data):
        """업종 PER 모델 기반 적정주가를 계산합니다."""
        indust_per = data.get('indust_per')
        eps_pred = data.get('eps_pred')
        if indust_per is not None and eps_pred is not None and indust_per > 0 and eps_pred > 0:
            return eps_pred * indust_per
        return 0

    def _calculate_fair_value_pegr(self, data, eps_growth_rate):
        """PEGR 모델 기반 적정주가를 계산합니다."""
        eps_pred = data.get('eps_pred')
        if eps_pred is not None and eps_pred > 0 and \
           eps_growth_rate is not None and self.PEGR_MIN_GROWTH < eps_growth_rate < self.PEGR_MAX_GROWTH:
            return eps_growth_rate * eps_pred
        return 0

    def _blend_and_apply_margin(self, fv_rim, fv_per, fv_pegr):
        """여러 모델의 적정주가를 혼합하고 안전마진을 적용합니다."""
        weighted_sum, total_weight = 0, 0
        
        if fv_rim > 0:
            weighted_sum += fv_rim * self.W_RIM
            total_weight += self.W_RIM
        if fv_per > 0:
            weighted_sum += fv_per * self.W_PER
            total_weight += self.W_PER
        if fv_pegr > 0:
            weighted_sum += fv_pegr * self.W_PEGR
            total_weight += self.W_PEGR
            
        if total_weight > 0:
            blended_value = weighted_sum / total_weight
            return blended_value * self.CONSERVATIVE_FACTOR
        return 0

    def _classify_valuation(self, current_price, fair_value, per, eps_growth_rate):
        """최종 적정주가를 바탕으로 종목을 분류하고 관련 지표를 계산합니다."""
        discrepancy_ratio = ((current_price - fair_value) / fair_value) * 100
        
        peg_ratio = None
        if per and per > 0 and eps_growth_rate and eps_growth_rate > self.PEGR_MIN_GROWTH:
            peg_ratio = per / eps_growth_rate

        if discrepancy_ratio < self.LOW_VALUATION_THRESHOLD:
            result = self._RESULT_UNDERVALUED
        elif discrepancy_ratio > self.HIGH_VALUATION_THRESHOLD:
            result = self._RESULT_OVERVALUED
        else:
            result = self._RESULT_FAIR
            
        return result, discrepancy_ratio, peg_ratio

    def _parse_perf_value(self, val):
        """실적 이슈 텍스트를 숫자로 변환합니다."""
        try:
            return float(str(val).replace(',', '').replace('%', '').strip())
        except (ValueError, TypeError, AttributeError):
            return val

    def _perform_valuation_calculation(self, stock_data):
        """주어진 데이터를 바탕으로 개별 종목의 가치를 평가합니다."""
        try:
            data = self._prepare_data(stock_data)
            growth_rates = self._calculate_growth_rates(data)

            fv_rim = self._calculate_fair_value_rim(data)
            fv_per = self._calculate_fair_value_per(data)
            fv_pegr = self._calculate_fair_value_pegr(data, growth_rates['eps_growth_rate'])

            fair_value = self._blend_and_apply_margin(fv_rim, fv_per, fv_pegr)
            if fair_value <= 0: return None

            result, discrepancy_ratio, peg_ratio = self._classify_valuation(
                data['current_price'], fair_value, data['per'], growth_rates['eps_growth_rate']
            )

            return {
                'code': data['code'], 'name': data['name'], 'current_price': data['current_price'],
                'fair_value': round(fair_value, 2),
                'discrepancy_ratio': round(discrepancy_ratio, 2),
                'pbr': data['pbr'], 'per': data['per'], 'roe': data['roe'],
                'eps_growth_rate': round(growth_rates['eps_growth_rate'], 2),
                'bps_growth_rate': round(growth_rates['bps_growth_rate'], 2),
                'roe_growth_rate': round(growth_rates['roe_growth_rate'], 2),
                'peg_ratio': round(peg_ratio, 2) if peg_ratio is not None else None,
                'result': result,
                'perf_yoy': self._parse_perf_value(data['perf_yoy']),
                'perf_vs_3m_ago': self._parse_perf_value(data['perf_vs_3m_ago']),
            }
        except Exception as e:
            logging.error(f"'{stock_data.get('code')}' 가치 평가 계산 중 오류: {e}")
            return None

    def _save_results_to_db(self, results, date_str):
        """평가 결과를 DB에 저장합니다."""
        try:
            query = """
            INSERT INTO valuations (
                code, date, fair_value, current_price, discrepancy_ratio,
                eps_growth_rate, peg_ratio, valuation_result
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                fair_value = VALUES(fair_value),
                current_price = VALUES(current_price),
                discrepancy_ratio = VALUES(discrepancy_ratio),
                eps_growth_rate = VALUES(eps_growth_rate),
                peg_ratio = VALUES(peg_ratio),
                valuation_result = VALUES(valuation_result)
            """
            params = [(r['code'], date_str, r['fair_value'], r['current_price'],
                       r['discrepancy_ratio'], r['eps_growth_rate'], r.get('peg_ratio'), r['result']) for r in results]
            
            self.db_access.execute_many_query(query, params)
            logging.info(f"가치 평가 결과를 DB에 저장했습니다. ({len(results)}건)")
        except Exception as e:
            logging.error(f"DB 저장 중 오류 발생: {e}")

    def _save_results_to_excel(self, results, date_str):
        """평가 결과를 Excel 파일로 저장합니다."""
        if not results:
            logging.info("저장할 가치 평가 결과가 없습니다.")
            return

        try:
            df = pd.DataFrame(results)
            
            # 컬럼 순서 조정: name을 code 바로 뒤로 이동
            if 'name' in df.columns:
                cols = df.columns.tolist()
                cols.insert(1, cols.pop(cols.index('name')))
                df = df[cols]

            report_dir = 'reports'
            os.makedirs(report_dir, exist_ok=True)
            filename = os.path.join(report_dir, f'valuation_results_{date_str}.xlsx')
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Valuation')
                worksheet = writer.sheets['Valuation']

                # 컬럼 인덱스 매핑
                col_map = {col: idx + 1 for idx, col in enumerate(df.columns)}

                # 0. Code 컬럼 텍스트 포맷 적용
                if 'code' in col_map:
                    code_col_idx = col_map['code']
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=code_col_idx)
                        cell.number_format = '@'

                # 1. 숫자 포맷 적용 (천 단위 콤마, 소수점 제거)
                for col_name in ['current_price', 'fair_value']:
                    if col_name in col_map:
                        col_idx = col_map[col_name]
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col_idx)
                            cell.number_format = '#,##0'

                # 1-1. 소수점 2자리 포맷 적용 (실적 지표)
                for col_name in ['perf_yoy', 'perf_vs_3m_ago']:
                    if col_name in col_map:
                        col_idx = col_map[col_name]
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col_idx)
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '0.00'

                # 2. '저평가' 종목 행 전체 배경색 적용
                if 'result' in col_map:
                    result_col_idx = col_map['result']
                    undervalued_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                    
                    for row_num in range(2, worksheet.max_row + 1):
                        result_cell = worksheet.cell(row=row_num, column=result_col_idx)
                        if result_cell.value == self._RESULT_UNDERVALUED:
                            for col_num in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row_num, column=col_num).fill = undervalued_fill

                # 3. 컬럼 너비 자동 조정
                for column_cells in worksheet.columns:
                    max_length = 0
                    col_letter = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if cell.value:
                                # 한글 등 멀티바이트 문자 고려하여 길이 계산 (한글은 길이를 더 크게 잡음)
                                cell_len = len(str(cell.value)) + sum(1 for c in str(cell.value) if ord(c) > 127)
                                if cell_len > max_length:
                                    max_length = cell_len
                        except:
                            pass
                    worksheet.column_dimensions[col_letter].width = max_length + 2

            logging.info(f"가치 평가 결과를 '{filename}' 파일로 성공적으로 저장했습니다.")
        except Exception as e:
            logging.error(f"Excel 파일 저장 중 오류 발생: {e}")

if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    with get_db_connection() as db:
        valuation_manager = ValuationManager(db)
        # 모든 종목에 대해 가치 평가 실행
        valuation_manager.calculate_and_save_valuations()

    logging.info("ValuationManager 스크립트 실행 완료.")
