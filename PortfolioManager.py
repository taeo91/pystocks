import pandas as pd
from DBAccessManager import DBAccessManager
import logging
import os
from dotenv import load_dotenv
import datetime
import re
import FinanceDataReader as fdr
from ETFManager import ETFManager
from AppManager import get_portfolio_excel_path
from openpyxl import load_workbook
from openpyxl.styles import Font

class PortfolioManager:
    def __init__(self, db_manager, etf_manager):
        self.db_manager = db_manager
        self.etf_manager = etf_manager
        self.target_sheets = ['CMA', '연금저축', 'IRP', '개인연금', '퇴직연금']
        self.sheet_configs = {}

    # --- Private Helper Methods ---

    def _get_market_lookup_tables(self):
        """Fetches and consolidates market and name lookup tables from KOSPI, KOSDAQ, and ETF listings."""
        name_lookup, market_lookup = {}, {}
        sources = {'KOSPI': 'KOSPI', 'KOSDAQ': 'KOSDAQ', 'ETF': 'ETF'}
        
        for name, source in sources.items():
            try:
                logging.info(f"Fetching {name} listings...")
                listing = fdr.StockListing(source)
                listing['Symbol'] = listing['Symbol'].astype(str)
                name_lookup.update(listing.set_index('Symbol')['Name'].to_dict())
                if source != 'ETF':
                    market_lookup.update(listing.set_index('Symbol')['Market'].to_dict())
                else:
                    for ticker in listing['Symbol']:
                        market_lookup[ticker] = 'ETF'
            except Exception as e:
                logging.warning(f"Could not fetch {name} listings: {e}")
        return name_lookup, market_lookup

    def _get_or_create_company_ids(self, tickers, name_lookup, market_lookup):
        """Ensures all tickers exist as companies in the DB and returns a {ticker: company_id} map."""
        if not tickers:
            return {}

        placeholders = ', '.join(['%s'] * len(tickers))
        query = f"SELECT code FROM companies WHERE code IN ({placeholders})"
        existing_rows = self.db_manager.fetch_all(query, tickers)
        existing_tickers = {row[0] for row in existing_rows} if existing_rows else set()
        new_tickers = [t for t in tickers if t not in existing_tickers]

        if new_tickers:
            logging.info(f"Found {len(new_tickers)} new tickers to insert.")
            new_company_data = [
                (ticker, name_lookup.get(ticker, ticker), market_lookup.get(ticker, 'UNKNOWN'))
                for ticker in new_tickers
            ]
            insert_query = "INSERT INTO companies (code, name, market) VALUES (%s, %s, %s)"
            self.db_manager.execute_many_query(insert_query, new_company_data)
            for ticker in new_tickers:
                if market_lookup.get(ticker) == 'ETF':
                    self.etf_manager.add_etf(ticker, name_lookup.get(ticker, ticker))

        query = f"SELECT code, id FROM companies WHERE code IN ({placeholders})"
        company_rows = self.db_manager.fetch_all(query, tickers)
        return {code: id for code, id in company_rows} if company_rows else {}

    def _create_table_from_dataframe(self, table_name, df, add_auto_increment_id=True):
        """Helper to generate and execute a CREATE TABLE query from a DataFrame."""
        create_table_query = f"CREATE TABLE IF NOT EXISTS {table_name} ("
        if add_auto_increment_id:
            create_table_query += "`id` INT AUTO_INCREMENT PRIMARY KEY, "

        for column, dtype in df.dtypes.items():
            mysql_type = self.get_mysql_data_type(dtype)
            if column == 'ticker':
                mysql_type = 'VARCHAR(255)'
            create_table_query += f"`{column}` {mysql_type}, "
        
        create_table_query = create_table_query.rstrip(", ") + ")"
        self.db_manager.execute_query(create_table_query)
        logging.info(f"Table '{table_name}' created or already exists.")

    def _update_revision_portfolio(self, workbook, latest_db_date, tickers, placeholders):
        """Updates the portfolio sheets for revision files."""
        try:
            # Get prices from both tables
            query_stock = f"SELECT c.code, p.close_price FROM prices p JOIN companies c ON p.company_id = c.id WHERE p.trade_date = %s AND c.code IN ({placeholders}) AND c.market != 'ETF'"
            stock_prices_raw = self.db_manager.fetch_all(query_stock, [latest_db_date] + tickers)
            prices_for_date = {code: price for code, price in stock_prices_raw} if stock_prices_raw else {}

            query_etf = f"SELECT ei.code, p.close_price FROM etf_prices p JOIN etf_info ei ON p.etf_id = ei.id WHERE p.trade_date = %s AND ei.code IN ({placeholders})"
            etf_prices_raw = self.db_manager.fetch_all(query_etf, [latest_db_date] + tickers)
            etf_prices = {code: price for code, price in etf_prices_raw} if etf_prices_raw else {}

            prices_for_date.update(etf_prices)
            blue_font = Font(color="0000FF")

            for sheet_name in self.target_sheets:
                if sheet_name not in workbook.sheetnames or sheet_name not in self.sheet_configs:
                    continue
                sheet = workbook[sheet_name]
                config = self.sheet_configs[sheet_name]
                
                if config['orientation'] == 'horizontal':
                    header_row_num = config['header_idx'] + 1
                    try:
                        header = [cell.value for cell in sheet[header_row_num]]
                        code_col_idx = header.index('코드') + 1
                        price_col_idx = header.index('현재가') + 1
                    except ValueError:
                        continue
                    for row_idx in range(header_row_num + 1, sheet.max_row + 1):
                        ticker_cell = sheet.cell(row=row_idx, column=code_col_idx)
                        ticker = str(ticker_cell.value).zfill(6) if ticker_cell.value else None
                        if ticker and ticker in prices_for_date:
                            price_cell = sheet.cell(row=row_idx, column=price_col_idx)
                            price_cell.value = prices_for_date[ticker]
                            price_cell.font = blue_font
                elif config['orientation'] == 'vertical':
                    header_col_num = config['header_idx'] + 1
                    try:
                        header = [sheet.cell(row=i, column=header_col_num).value for i in range(1, sheet.max_row + 1)]
                        code_row_idx = header.index('코드') + 1
                        price_row_idx = header.index('현재가') + 1
                    except ValueError:
                        continue
                    for col_idx in range(header_col_num + 1, sheet.max_column + 1):
                        ticker_cell = sheet.cell(row=code_row_idx, column=col_idx)
                        ticker = str(ticker_cell.value).zfill(6) if ticker_cell.value else None
                        if ticker and ticker in prices_for_date:
                            price_cell = sheet.cell(row=price_row_idx, column=col_idx)
                            price_cell.value = prices_for_date[ticker]
                            price_cell.font = blue_font

        except (KeyError, Exception) as e:
            logging.error(f"Failed to update sheets: {e}")

    # --- Public Methods ---

    def get_mysql_data_type(self, dtype):
        if "int" in str(dtype): return "INT"
        if "float" in str(dtype): return "FLOAT"
        if "datetime" in str(dtype): return "DATETIME"
        return "VARCHAR(255)"

    def get_tickers_from_excel(self, file_path):
        """
        Dynamically finds the header in target sheets (row or column) and extracts stock tickers.
        """
        all_tickers = set()
        try:
            xls = pd.ExcelFile(file_path)
            for sheet_name in self.target_sheets:
                if sheet_name not in xls.sheet_names:
                    logging.warning(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
                    continue

                df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                found_header = False

                # 1. Try to find header in rows (horizontal orientation)
                for i, row in df_full.head(10).iterrows():
                    row_values = [str(v).strip() for v in row.dropna().tolist()]
                    if '코드' in row_values and '종목' in row_values:
                        self.sheet_configs[sheet_name] = {'header_idx': i, 'orientation': 'horizontal'}
                        df = pd.read_excel(file_path, sheet_name=sheet_name, header=i, dtype={'코드': str})
                        if '코드' in df.columns:
                            tickers = df['코드'].dropna().astype(str).apply(lambda x: x.zfill(6)).tolist()
                            all_tickers.update(tickers)
                            logging.info(f"'{file_path}' 파일의 '{sheet_name}' 시트에서 수평 방향으로 {len(tickers)}개의 티커를 추출했습니다. (헤더 행: {i + 1})")
                        found_header = True
                        break

                if found_header: continue

                # 2. If not found, try to find header in columns (vertical orientation)
                df_transposed = df_full.T
                for i, row in df_transposed.head(10).iterrows():
                    row_values = [str(v).strip() for v in row.dropna().tolist()]
                    if '코드' in row_values and '종목' in row_values:
                        self.sheet_configs[sheet_name] = {'header_idx': i, 'orientation': 'vertical'}
                        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None).T
                        df.columns = df.iloc[i]
                        df = df.iloc[i+1:]
                        df = df.rename(columns={'코드': '코드_val'})
                        
                        if '코드_val' in df.columns:
                            tickers = df['코드_val'].dropna().astype(str).apply(lambda x: x.zfill(6)).tolist()
                            all_tickers.update(tickers)
                            logging.info(f"'{file_path}' 파일의 '{sheet_name}' 시트에서 수직 방향으로 {len(tickers)}개의 티커를 추출했습니다. (헤더 열: {i + 1})")
                        found_header = True
                        break

                if not found_header:
                    logging.error(f"'{sheet_name}' 시트에서 '코드'와 '종목'을 포함하는 헤더를 찾을 수 없습니다.")

            return list(all_tickers)
            
        except Exception as e:
            logging.error(f"Excel 파일 '{file_path}'에서 종목 정보를 읽는 중 오류 발생: {e}", exc_info=True)
            return []

    def fetch_and_save_prices(self, tickers, market_type='etf', start_date=None):
        default_start_date = (datetime.date.today() - datetime.timedelta(days=365)).strftime('%Y-%m-%d')
        fetch_start_date_overall = start_date if start_date else default_start_date

        is_etf = market_type.lower() == 'etf'
        table_name = 'etf_prices' if is_etf else 'prices'
        id_column = 'etf_id' if is_etf else 'company_id'

        name_lookup, market_lookup = self._get_market_lookup_tables()

        # 1. Filter tickers for the given market type
        if is_etf:
            target_tickers = [t for t in tickers if market_lookup.get(t) == 'ETF']
        else: # For 'stock', get everything that is not an ETF.
            target_tickers = [t for t in tickers if market_lookup.get(t) != 'ETF']

        if not target_tickers:
            logging.info(f"No tickers to process for market type '{market_type}'.")
            return

        # 2. Ensure all tickers exist in DB and get their respective IDs
        self._get_or_create_company_ids(target_tickers, name_lookup, market_lookup)

        placeholders = ', '.join(['%s'] * len(target_tickers))
        id_map_query = f"SELECT code, id FROM {'etf_info' if is_etf else 'companies'} WHERE code IN ({placeholders})"
        id_map_rows = self.db_manager.fetch_all(id_map_query, target_tickers)
        id_map = {code: id for code, id in id_map_rows} if id_map_rows else {}

        # 3. Find the last trade date for each ID to fetch only new data
        db_ids = list(id_map.values())
        last_trade_dates = {}
        if db_ids:
            placeholders_ids = ', '.join(['%s'] * len(db_ids))
            query = f"SELECT {id_column}, MAX(trade_date) FROM {table_name} WHERE {id_column} IN ({placeholders_ids}) GROUP BY {id_column}"
            last_dates_raw = self.db_manager.fetch_all(query, db_ids)
            last_trade_dates = {cid: date for cid, date in last_dates_raw} if last_dates_raw else {}

        # 4. Fetch and save price data for each ticker
        for ticker in target_tickers:
            db_id = id_map.get(ticker)
            if not db_id:
                logging.warning(f"Could not process ticker {ticker}, no company_id found. Skipping.")
                continue

            last_date = last_trade_dates.get(db_id)
            fetch_start_date = (last_date + datetime.timedelta(days=1)).strftime('%Y-%m-%d') if last_date else fetch_start_date_overall
            
            logging.info(f"[{ticker}] Fetching price data from {fetch_start_date}")
            try:
                df = fdr.DataReader(ticker, start=fetch_start_date)
                if df.empty:
                    data_to_insert = [(db_id, date.strftime('%Y-%m-%d'), row['Open'], row['High'], row['Low'], row['Close'], row['Volume']) for date, row in df.iterrows()]
                if data_to_insert:
                    price_query = f"INSERT INTO {table_name} ({id_column}, trade_date, open_price, high_price, low_price, close_price, volume) VALUES (%s, %s, %s, %s, %s, %s, %s) ON DUPLICATE KEY UPDATE open_price=VALUES(open_price), high_price=VALUES(high_price), low_price=VALUES(low_price), close_price=VALUES(close_price), volume=VALUES(volume)"
                    self.db_manager.execute_many_query(price_query, data_to_insert)
                    logging.info(f"[{ticker}] Inserted/updated {len(data_to_insert)} price records into {table_name}.")
                else:
                    logging.info(f"[{ticker}] No new price data found.")
            except Exception as e:
                logging.error(f"Error fetching or saving price data for {ticker}: {e}")

    def import_portfolio_from_excel(self, file_path, table_name):
        try:
            all_dfs = []
            for sheet_name in self.target_sheets:
                if sheet_name not in self.sheet_configs:
                    continue
                config = self.sheet_configs[sheet_name]
                if config['orientation'] == 'horizontal':
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=config['header_idx'])
                elif config['orientation'] == 'vertical':
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None).T
                    df.columns = df.iloc[config['header_idx']]
                    df = df.iloc[config['header_idx']+1:]
                else:
                    continue
                all_dfs.append(df)
            
            if not all_dfs:
                logging.error("Import할 수 있는 포트폴리오 데이터가 없습니다.")
                return
                
            df = pd.concat(all_dfs, ignore_index=True)
            logging.info(f"Successfully read Excel file: {file_path}")
        except FileNotFoundError:
            logging.error(f"Error: The file '{file_path}' was not found.")
            return

        df.columns = [str(col).strip().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "") for col in df.columns]
        self._create_table_from_dataframe(table_name, df, add_auto_increment_id=False)

        insert_query = f"INSERT INTO {table_name} ({', '.join([f'`{col}`' for col in df.columns])}) VALUES ({', '.join(['%s'] * len(df.columns))})"
        df = df.where(pd.notna(df), None)
        data_to_insert = [tuple(row) for row in df.to_numpy()]
        self.db_manager.execute_many_query(insert_query, data_to_insert)
        logging.info(f"Data inserted into '{table_name}'.")

    def import_settings_from_excel(self, file_path, table_name):
        sheet_name = '설정'
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            settings_df = pd.DataFrame({
                '날짜': [df.iloc[2, 1]], 'CMA순자산평가금': [df.iloc[4, 1]],
                '추정현금': [df.iloc[5, 1]], '현금목표비중': [df.iloc[7, 1]]
            })
        except (FileNotFoundError, IndexError, Exception) as e:
            logging.error(f"Error reading or processing sheet '{sheet_name}': {e}")
            return
        
        settings_df.columns = [col.strip().replace(" ", "_") for col in settings_df.columns]
        self._create_table_from_dataframe(table_name, settings_df)

        self.db_manager.execute_query(f"TRUNCATE TABLE {table_name}")
        insert_query = f"INSERT INTO {table_name} ({', '.join([f'`{col}`' for col in settings_df.columns])}) VALUES ({', '.join(['%s'] * len(settings_df.columns))})"
        data_to_insert = [tuple(row) for row in settings_df.to_numpy()]
        self.db_manager.execute_many_query(insert_query, data_to_insert)
        logging.info(f"Data inserted into '{table_name}'.")

    def import_holdings_from_excel(self, file_path, table_name):
        """Imports holdings data from the target sheets into a MySQL database."""
        cols_map = {
            '코드': 'ticker', '종목명': 'stock_name', '보유수량': 'quantity',
            '매입단가': 'avg_price', '매입금액': 'purchase_amount', 'MDD': 'mdd_percent'
        }

        try:
            all_dfs = []
            for sheet_name in self.target_sheets:
                if sheet_name not in self.sheet_configs:
                    continue
                config = self.sheet_configs[sheet_name]
                if config['orientation'] == 'horizontal':
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=config['header_idx'], dtype={'코드': str})
                elif config['orientation'] == 'vertical':
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None).T
                    df.columns = df.iloc[config['header_idx']]
                    df = df.iloc[config['header_idx']+1:]
                    df = df.astype({'코드': str})
                else:
                    continue
                all_dfs.append(df)
            
            if not all_dfs:
                logging.error("Import할 수 있는 Holdings 데이터가 없습니다.")
                return
                
            df = pd.concat(all_dfs, ignore_index=True)
            holdings_df = df[list(cols_map.keys())].rename(columns=cols_map)
        except (FileNotFoundError, KeyError, Exception) as e:
            logging.error(f"Error reading or processing sheets: {e}")
            return

        holdings_df = holdings_df[holdings_df['stock_name'] != '합계'].dropna(subset=['stock_name'])
        if 'ticker' in holdings_df.columns and holdings_df['ticker'].notna().any():
            holdings_df['ticker'] = holdings_df['ticker'].dropna().apply(lambda x: str(x).zfill(6))

        ticker_list = holdings_df['ticker'].dropna().tolist()
        if ticker_list:
            placeholders = ', '.join(['%s'] * len(ticker_list))
            
            # Query for stock prices
            query_stocks = f"""
                SELECT c.code, p.close_price FROM prices p
                JOIN (
                    SELECT company_id, MAX(trade_date) as max_date FROM prices
                    WHERE company_id IN (SELECT id FROM companies WHERE code IN ({placeholders}))
                    GROUP BY company_id
                ) latest ON p.company_id = latest.company_id AND p.trade_date = latest.max_date
                JOIN companies c ON p.company_id = c.id
            """
            stock_prices_raw = self.db_manager.fetch_all(query_stocks, ticker_list)
            price_map = {code: price for code, price in stock_prices_raw} if stock_prices_raw else {}
            
            # Query for ETF prices
            query_etfs = f"""
                SELECT ei.code, p.close_price FROM etf_prices p
                JOIN (
                    SELECT etf_id, MAX(trade_date) as max_date FROM etf_prices
                    WHERE etf_id IN (SELECT id FROM etf_info WHERE code IN ({placeholders}))
                    GROUP BY etf_id
                ) latest ON p.etf_id = latest.etf_id AND p.trade_date = latest.max_date
                JOIN etf_info ei ON p.etf_id = ei.id
            """
            etf_prices_raw = self.db_manager.fetch_all(query_etfs, ticker_list)
            etf_price_map = {code: price for code, price in etf_prices_raw} if etf_prices_raw else {}

            price_map.update(etf_price_map)

            holdings_df['close_price'] = holdings_df['ticker'].map(price_map)
        else:
            holdings_df['close_price'] = None

        self._create_table_from_dataframe(table_name, holdings_df)
        self.db_manager.execute_query(f"TRUNCATE TABLE {table_name}")

        insert_query = f"INSERT INTO {table_name} ({', '.join([f'`{col}`' for col in holdings_df.columns])}) VALUES ({', '.join(['%s'] * len(holdings_df.columns))})"
        holdings_df = holdings_df.where(pd.notna(holdings_df), None)
        data_to_insert = [tuple(row) for row in holdings_df.to_numpy()]
        self.db_manager.execute_many_query(insert_query, data_to_insert)
        logging.info(f"Data inserted into '{table_name}'.")

    def update_portfolio_excel_with_prices(self, file_path):
        """Updates the 'Portfolio' sheet with the latest stock prices from the database."""
        try:
            logging.info("--- Starting Excel update process ---")
            workbook = load_workbook(file_path)
            all_tickers = self.get_tickers_from_excel(file_path)
            if not all_tickers:
                logging.warning("No tickers found. Aborting update.")
                return

            placeholders = ', '.join(['%s'] * len(all_tickers))
            
            # Get latest date from both tables
            query_stock = f"SELECT MAX(p.trade_date) FROM prices p JOIN companies c ON p.company_id = c.id WHERE c.code IN ({placeholders})"
            result_stock = self.db_manager.fetch_one(query_stock, all_tickers)
            latest_stock_date = result_stock[0] if result_stock and result_stock[0] else None

            query_etf = f"SELECT MAX(p.trade_date) FROM etf_prices p JOIN etf_info ei ON p.etf_id = ei.id WHERE ei.code IN ({placeholders})"
            result_etf = self.db_manager.fetch_one(query_etf, all_tickers)
            latest_etf_date = result_etf[0] if result_etf and result_etf[0] else None

            latest_db_date = None
            if latest_stock_date and latest_etf_date:
                latest_db_date = max(latest_stock_date, latest_etf_date)
            elif latest_stock_date:
                latest_db_date = latest_stock_date
            else:
                latest_db_date = latest_etf_date
            
            if not latest_db_date:
                logging.warning("No price data in DB for any tickers. Aborting.")
                return
                
            try:
                settings_sheet = workbook['설정']
                cell_to_update = settings_sheet['B3']
                cell_to_update.value = latest_db_date
                cell_to_update.font = Font(color="0000FF")
                logging.info(f"Updated '설정' sheet with latest date: {latest_db_date}")
            except (KeyError, Exception) as e:
                logging.warning(f"Could not update '설정' sheet: {e}")

            self._update_revision_portfolio(workbook, latest_db_date, all_tickers, placeholders)

            workbook.save(file_path)
            logging.info("--- Excel file update process finished ---")
        except FileNotFoundError:
            logging.error(f"Error: File '{file_path}' not found.")
        except Exception as e:
            logging.error(f"An error occurred while updating excel: {e}", exc_info=True)
if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    db_manager = DBAccessManager(
        host=os.getenv("DB_HOST"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        database=os.getenv("DB_NAME")
    )
    
    portfolio_file_path = get_portfolio_excel_path()
    if not portfolio_file_path:
        logging.error("포트폴리오 파일 경로가 없어서 PortfolioManager를 실행할 수 없습니다.")
    else:
        db_manager.connect_to_mysql()

        if db_manager.connection:
            etf_manager = ETFManager(db_manager)
            portfolio_manager = PortfolioManager(db_manager, etf_manager)
            
            logging.info("Step 1: Reading tickers from Excel file...")
            tickers = portfolio_manager.get_tickers_from_excel(portfolio_file_path)

            if tickers:
                logging.info("Step 2: Fetching latest prices and saving to DB...")
                portfolio_manager.fetch_and_save_prices(tickers, market_type='etf')
                portfolio_manager.fetch_and_save_prices(tickers, market_type='stock')

                logging.info("Step 3: Updating Excel file with latest prices from DB...")
                portfolio_manager.update_portfolio_excel_with_prices(portfolio_file_path)
            
            db_manager.close_connection()
