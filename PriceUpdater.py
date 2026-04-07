
import logging
import os
from dotenv import load_dotenv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import FinanceDataReader as fdr
import datetime

# --- Functions for Price Updates ---

def get_tickers_from_excel(file_path, sheet_name):
    """
    Dynamically finds the header in the given sheet (row or column) and extracts stock tickers.
    Returns a list of tickers, the header index, and the orientation ('horizontal' or 'vertical').
    """
    try:
        xls = pd.ExcelFile(file_path)
        if sheet_name not in xls.sheet_names:
            logging.error(f"'{sheet_name}' 시트를 찾을 수 없습니다. 사용 가능한 시트: {xls.sheet_names}")
            return [], None, None

        df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # 1. Try to find header in rows (horizontal orientation)
        for i, row in df_full.head(10).iterrows():
            row_values = [str(v).strip() for v in row.dropna().tolist()]
            if '코드' in row_values and '종목' in row_values:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=i, dtype={'코드': str})
                if '코드' in df.columns:
                    tickers = df['코드'].dropna().astype(str).apply(lambda x: x.zfill(6)).tolist()
                    logging.info(f"'{file_path}' 파일의 '{sheet_name}' 시트에서 수평 방향으로 {len(tickers)}개의 티커를 추출했습니다. (헤더 행: {i + 1})")
                    return tickers, i, 'horizontal'

        # 2. If not found, try to find header in columns (vertical orientation)
        df_transposed = df_full.T
        for i, row in df_transposed.head(10).iterrows():
            row_values = [str(v).strip() for v in row.dropna().tolist()]
            if '코드' in row_values and '종목' in row_values:
                # Header is found in a column, so we use the transposed dataframe
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None).T
                df.columns = df.iloc[i] # Set the found header column as header
                df = df.iloc[i+1:] # Drop header row
                df = df.rename(columns={'코드': '코드_val'}) # Rename to avoid conflict
                
                if '코드_val' in df.columns:
                    tickers = df['코드_val'].dropna().astype(str).apply(lambda x: x.zfill(6)).tolist()
                    logging.info(f"'{file_path}' 파일의 '{sheet_name}' 시트에서 수직 방향으로 {len(tickers)}개의 티커를 추출했습니다. (헤더 열: {i + 1})")
                    return tickers, i, 'vertical'

        logging.error(f"'{sheet_name}' 시트에서 '코드'와 '종목'을 포함하는 헤더를 찾을 수 없습니다.")
        return [], None, None

    except Exception as e:
        logging.error(f"Excel 파일 '{file_path}'의 '{sheet_name}' 시트에서 종목 정보를 읽는 중 오류 발생: {e}", exc_info=True)
        return [], None, None


def get_current_prices(tickers):
    """Fetches the current price for a list of tickers."""
    prices = {}
    # Fetch data for a recent period (e.g., last 10 calendar days) to ensure we get the latest available trading day's price.
    # This handles weekends and holidays automatically by taking the last available entry.
    start_date_for_fetch = (datetime.date.today() - datetime.timedelta(days=10)).strftime('%Y-%m-%d')

    for ticker in tickers:
        try:
            df = fdr.DataReader(ticker, start=start_date_for_fetch)
            if not df.empty:
                current_price = df.iloc[-1]['Close']
                prices[ticker] = current_price
                logging.info(f"[{ticker}] 최신 현재가 조회 완료: {current_price}")
            else:
                logging.warning(f"[{ticker}] 최근 10일간의 현재가 정보를 가져올 수 없습니다.")
        except Exception as e:
            logging.error(f"'{ticker}'의 가격 조회 중 오류 발생: {e}")
    return prices


# --- Function for Excel Update ---

def update_excel_file(file_path, sheet_updates):
    """Updates the specified sheets with the latest stock prices based on orientation."""
    if not sheet_updates:
        logging.warning("업데이트할 시트 정보가 없습니다. Excel 파일 업데이트를 건너뜁니다.")
        return

    try:
        logging.info("--- Excel 파일 업데이트 시작 ---")
        workbook = load_workbook(file_path)
        blue_font = Font(color="0000FF")

        for sheet_name, data in sheet_updates.items():
            prices = data.get('prices')
            header_idx = data.get('header_idx')
            orientation = data.get('orientation')

            if not prices:
                logging.warning(f"'{sheet_name}' 시트에 업데이트할 가격 정보가 없습니다.")
                continue
            
            if header_idx is None or orientation is None:
                logging.error(f"헤더 정보나 방향이 없어 '{sheet_name}' 시트를 업데이트할 수 없습니다.")
                continue

            if sheet_name not in workbook.sheetnames:
                logging.error(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
                continue

            sheet = workbook[sheet_name]
            logging.info(f"'{sheet_name}' 시트 업데이트 중 (방향: {orientation})")

            if orientation == 'horizontal':
                header_row_num = header_idx + 1
                try:
                    header = [cell.value for cell in sheet[header_row_num]]
                    code_col_idx = header.index('코드') + 1
                    price_col_idx = header.index('현재가') + 1
                except (ValueError, IndexError):
                    logging.error(f"'{sheet_name}' 시트의 {header_row_num}행에서 '코드' 또는 '현재가' 헤더를 찾을 수 없습니다.")
                    continue

                for row_idx in range(header_row_num + 1, sheet.max_row + 1):
                    ticker_cell = sheet.cell(row=row_idx, column=code_col_idx)
                    ticker = str(ticker_cell.value).zfill(6) if ticker_cell.value else None
                    if ticker and ticker in prices:
                        price_cell = sheet.cell(row=row_idx, column=price_col_idx)
                        price_cell.value = prices[ticker]
                        price_cell.font = blue_font
                        logging.info(f"'{sheet_name}' 시트: '{ticker}'의 현재가를 {prices[ticker]}로 업데이트했습니다.")

            elif orientation == 'vertical':
                header_col_num = header_idx + 1
                try:
                    header = [sheet.cell(row=i, column=header_col_num).value for i in range(1, sheet.max_row + 1)]
                    code_row_idx = header.index('코드') + 1
                    price_row_idx = header.index('현재가') + 1
                except (ValueError, IndexError):
                    logging.error(f"'{sheet_name}' 시트의 {header_col_num}열에서 '코드' 또는 '현재가' 헤더를 찾을 수 없습니다.")
                    continue

                for col_idx in range(header_col_num + 1, sheet.max_column + 1):
                    ticker_cell = sheet.cell(row=code_row_idx, column=col_idx)
                    ticker = str(ticker_cell.value).zfill(6) if ticker_cell.value else None
                    if ticker and ticker in prices:
                        price_cell = sheet.cell(row=price_row_idx, column=col_idx)
                        price_cell.value = prices[ticker]
                        price_cell.font = blue_font
                        logging.info(f"'{sheet_name}' 시트: '{ticker}'의 현재가를 {prices[ticker]}로 업데이트했습니다.")
        
        if '설정' in workbook.sheetnames:
            settings_sheet = workbook['설정']
            cell_to_update = settings_sheet['B3'] 
            today_date = datetime.date.today()
            cell_to_update.value = today_date
            cell_to_update.font = Font(color="0000FF")
            logging.info(f"'설정' 시트의 날짜를 {today_date.strftime('%Y-%m-%d')}로 업데이트했습니다.")

        workbook.save(file_path)
        logging.info("--- Excel 파일 업데이트 완료 ---")

    except FileNotFoundError:
        logging.error(f"오류: 파일 '{file_path}'을(를) 찾을 수 없습니다.")
    except Exception as e:
        logging.error(f"Excel 업데이트 중 오류 발생: {e}", exc_info=True)


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    load_dotenv()
    
    portfolio_file_path = os.getenv("PORTFOLIO_EXCEL_FILE", "reports/portfolio_r16.xlsx")
    
    if not portfolio_file_path or not os.path.exists(portfolio_file_path):
        logging.error(f"포트폴리오 파일이 잘못 설정되었거나 없습니다: {portfolio_file_path}")
    else:
        target_sheets = ['CMA', '연금저축', 'IRP']
        sheet_updates = {}
        all_tickers = set()

        for sheet in target_sheets:
            tickers, header_info, orientation = get_tickers_from_excel(portfolio_file_path, sheet)
            if tickers:
                sheet_updates[sheet] = {
                    'tickers': tickers,
                    'header_idx': header_info,
                    'orientation': orientation
                }
                all_tickers.update(tickers)
        
        if all_tickers:
            logging.info(f"=== 실시간 현재가 조회 시작 (총 {len(all_tickers)}개 종목) ===")
            current_prices = get_current_prices(list(all_tickers))
            
            for sheet in sheet_updates:
                sheet_updates[sheet]['prices'] = current_prices

            update_excel_file(portfolio_file_path, sheet_updates)
