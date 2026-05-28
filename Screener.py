import logging
import pandas as pd
import datetime
import os
import FinanceDataReader as fdr
from AppManager import get_db_connection
from openpyxl.styles import PatternFill, Font
from sqlalchemy import create_engine

class Screener:
    """
    특정 조건에 맞는 종목을 검색하고 결과를 출력하는 클래스
    """
    def __init__(self, db_access):
        """
        Screener 생성자

        Args:
            db_access (dbaccess): 데이터베이스 접근을 위한 dbaccess 객체
        """
        self.db_access = db_access
        # pandas 경고를 해결하기 위해 SQLAlchemy 엔진 생성
        db_info = {
            'user': os.getenv('DB_USER'),
            'password': os.getenv('DB_PASSWORD'),
            'host': os.getenv('DB_HOST'),
            'database': os.getenv('DB_NAME')
        }
        self.engine = create_engine(f"mysql+mysqlconnector://{db_info['user']}:{db_info['password']}@{db_info['host']}/{db_info['database']}")

    def find_leading_stocks(self, top_n=50, min_marcap_억=500, require_ma_alignment=False):
        """
        주도주를 스크리닝하여 Excel로 저장합니다.

        판별 기준:
        - RS Rating: 1M·3M·6M KOSPI 초과수익률 복합점수의 전체 백분위 (0~100)
        - 이평선 정배열: 현재가 > 20MA > 60MA > 120MA
        - 52주 신고가 근접도: 현재가 / 52주 최고가 (%)
        - 거래량 급증비: 최근 20일 평균 / 60일 평균 거래량

        Args:
            top_n: 출력할 상위 종목 수
            min_marcap_억: 최소 시가총액 필터 (억 원), None이면 무제한
            require_ma_alignment: True이면 이평선 정배열 종목만 출력
        """
        today = datetime.date.today()
        date_str = today.strftime('%Y-%m-%d')
        # 52주 신고가(252 거래일) + 120MA + 6M 여유분 → 약 400 calendar days
        start_date = (today - datetime.timedelta(days=400)).strftime('%Y-%m-%d')

        # 1. KOSPI 수익률 기준값
        logging.info("KOSPI 데이터를 가져옵니다...")
        try:
            kospi = fdr.DataReader('KS11', start=start_date)
            kospi_close = kospi['Close']
            kospi_close.index = pd.to_datetime(kospi_close.index).normalize()
        except Exception as e:
            logging.error(f"KOSPI 데이터 로드 실패: {e}")
            return None

        def kospi_return_pct(n_calendar_days):
            target = pd.Timestamp(today - datetime.timedelta(days=n_calendar_days))
            avail = kospi_close.index[kospi_close.index <= target]
            if avail.empty:
                return 0.0
            past = float(kospi_close.loc[avail[-1]])
            now = float(kospi_close.iloc[-1])
            return (now - past) / past * 100 if past > 0 else 0.0

        k_1m = kospi_return_pct(30)
        k_3m = kospi_return_pct(91)
        k_6m = kospi_return_pct(182)
        logging.info(f"KOSPI 수익률 — 1M: {k_1m:.2f}%, 3M: {k_3m:.2f}%, 6M: {k_6m:.2f}%")

        # 2. 전 종목 가격 데이터 (bulk fetch)
        logging.info("전 종목 가격 데이터를 가져옵니다...")
        price_query = """
            SELECT c.code, c.name, p.trade_date, p.close_price, p.volume
            FROM prices p
            JOIN companies c ON p.company_id = c.id
            WHERE p.trade_date >= %s
            ORDER BY c.code, p.trade_date
        """
        price_df = pd.read_sql(price_query, self.engine, params=(start_date,))
        if price_df.empty:
            logging.warning("가격 데이터가 없습니다.")
            return None
        price_df['trade_date'] = pd.to_datetime(price_df['trade_date']).dt.normalize()

        # code → name 매핑
        name_map = price_df.drop_duplicates('code').set_index('code')['name'].to_dict()

        # 피벗: 날짜 × 종목
        close_pivot = price_df.pivot_table(index='trade_date', columns='code', values='close_price')
        vol_pivot   = price_df.pivot_table(index='trade_date', columns='code', values='volume')

        # 3. 최신 재무 데이터 (marcap, beta)
        fin_query = """
            SELECT df.code, df.marcap, df.beta
            FROM daily_financials df
            JOIN (
                SELECT code, MAX(date) AS max_date FROM daily_financials GROUP BY code
            ) ldf ON df.code = ldf.code AND df.date = ldf.max_date
        """
        fin_df = pd.read_sql(fin_query, self.engine)

        # 4. 기준일 가격 추출 헬퍼
        def price_n_days_ago(pivot, n_calendar_days):
            target = pd.Timestamp(today - datetime.timedelta(days=n_calendar_days))
            avail = pivot.index[pivot.index <= target]
            return pivot.loc[avail[-1]] if not avail.empty else None

        price_now = close_pivot.iloc[-1]
        price_1m  = price_n_days_ago(close_pivot, 30)
        price_3m  = price_n_days_ago(close_pivot, 91)
        price_6m  = price_n_days_ago(close_pivot, 182)

        # 5. 종목별 지표 계산
        logging.info("종목별 지표를 계산합니다...")
        results = []
        for code in close_pivot.columns:
            prices = close_pivot[code].dropna()
            if len(prices) < 30:
                continue

            p_now = price_now.get(code)
            if not p_now or p_now <= 0:
                continue

            def ret_pct(p_past):
                v = p_past.get(code) if p_past is not None else None
                return (p_now - v) / v * 100 if v and v > 0 else None

            ret_1m = ret_pct(price_1m)
            ret_3m = ret_pct(price_3m)
            ret_6m = ret_pct(price_6m)

            rs_1m = (ret_1m - k_1m) if ret_1m is not None else None
            rs_3m = (ret_3m - k_3m) if ret_3m is not None else None
            rs_6m = (ret_6m - k_6m) if ret_6m is not None else None

            # 복합 RS: 1M 40% + 3M 35% + 6M 25%
            weighted = [(rs_1m, 0.40), (rs_3m, 0.35), (rs_6m, 0.25)]
            valid = [(v, w) for v, w in weighted if v is not None]
            if not valid:
                continue
            rs_composite = sum(v * w for v, w in valid) / sum(w for _, w in valid)

            # 이평선 정배열
            n = len(prices)
            ma20  = float(prices.tail(20).mean())  if n >= 20  else None
            ma60  = float(prices.tail(60).mean())  if n >= 60  else None
            ma120 = float(prices.tail(120).mean()) if n >= 120 else None
            ma_aligned = bool(
                ma20 and ma60 and ma120 and
                p_now > ma20 > ma60 > ma120
            )

            # 52주 신고가 근접도
            high_52w = float(prices.tail(252).max())
            proximity_52w = round(p_now / high_52w * 100, 1) if high_52w > 0 else None

            # 거래량 급증비
            vols = vol_pivot[code].dropna()
            vol_surge = None
            if len(vols) >= 60:
                avg20 = vols.tail(20).mean()
                avg60 = vols.tail(60).mean()
                vol_surge = round(float(avg20 / avg60), 2) if avg60 > 0 else None

            results.append({
                'code':           code,
                'name':           name_map.get(code, code),
                'current_price':  round(p_now),
                'ret_1m':         round(ret_1m, 2) if ret_1m is not None else None,
                'ret_3m':         round(ret_3m, 2) if ret_3m is not None else None,
                'ret_6m':         round(ret_6m, 2) if ret_6m is not None else None,
                'rs_1m':          round(rs_1m,  2) if rs_1m  is not None else None,
                'rs_3m':          round(rs_3m,  2) if rs_3m  is not None else None,
                'rs_6m':          round(rs_6m,  2) if rs_6m  is not None else None,
                'rs_composite':   round(rs_composite, 2),
                'ma_aligned':     ma_aligned,
                'proximity_52w':  proximity_52w,
                'vol_surge':      vol_surge,
            })

        if not results:
            logging.warning("스크리닝 결과가 없습니다.")
            return None

        result_df = pd.DataFrame(results)

        # RS Rating: 전체 종목 대비 백분위 (1~100)
        result_df['rs_rating'] = result_df['rs_composite'].rank(pct=True).mul(100).round(1)

        # 재무 데이터 병합
        result_df = result_df.merge(fin_df[['code', 'marcap', 'beta']], on='code', how='left')

        # 시가총액 필터
        if min_marcap_억:
            min_won = min_marcap_억 * 100_000_000
            result_df = result_df[result_df['marcap'].fillna(0) >= min_won]

        # 이평선 정배열 필터
        if require_ma_alignment:
            result_df = result_df[result_df['ma_aligned']]

        # RS Rating 기준 정렬
        result_df = result_df.sort_values('rs_rating', ascending=False).head(top_n)

        # marcap을 억 단위로 변환
        if 'marcap' in result_df.columns:
            result_df['marcap_억'] = (result_df['marcap'] / 1e8).round(0).astype('Int64')
            result_df = result_df.drop(columns=['marcap'])

        # rs_rating을 앞쪽으로 이동
        cols = result_df.columns.tolist()
        for key in ['rs_rating', 'rs_composite']:
            if key in cols:
                cols.insert(2, cols.pop(cols.index(key)))
        result_df = result_df[cols]

        # 6. Excel 저장
        os.makedirs('reports', exist_ok=True)
        filename = os.path.join('reports', f'leading_stocks_{date_str}.xlsx')
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='주도주')
            ws = writer.sheets['주도주']
            col_map = {col: i + 1 for i, col in enumerate(result_df.columns)}

            green_fill  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            bold_font   = Font(bold=True)

            for row_num in range(2, ws.max_row + 1):
                # 이평선 정배열 → 초록 배경
                if 'ma_aligned' in col_map:
                    if ws.cell(row=row_num, column=col_map['ma_aligned']).value is True:
                        for col_num in range(1, ws.max_column + 1):
                            ws.cell(row=row_num, column=col_map['ma_aligned']).fill = green_fill
                # RS Rating 90 이상 → bold
                if 'rs_rating' in col_map:
                    cell = ws.cell(row=row_num, column=col_map['rs_rating'])
                    if cell.value and float(cell.value) >= 90:
                        for col_num in range(1, ws.max_column + 1):
                            ws.cell(row=row_num, column=col_num).font = bold_font

            # 컬럼 너비 자동 조정
            for col_cells in ws.columns:
                max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
                ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2

        logging.info(f"주도주 스크리닝 완료: {len(result_df)}개 종목 → '{filename}'")
        return result_df

    def calculate_risk_metrics(self, code, months=3):
        """
        특정 종목의 최근 N개월간 리스크 지표(MDD, 최대 낙폭, 하락일 평균 하락폭)를 계산합니다.
        """
        try:
            start_date = datetime.date.today() - datetime.timedelta(days=months * 30)
            start_date_str = start_date.strftime('%Y-%m-%d')

            query = """
            WITH PeriodData AS (
                SELECT 
                    p.trade_date,
                    p.close_price,
                    p.high_price,
                    p.low_price,
                    LAG(p.close_price) OVER (ORDER BY p.trade_date) as prev_close,
                    -- 기간 내 누적 최고가 (MDD 계산용)
                    MAX(p.high_price) OVER (ORDER BY p.trade_date) as running_max_high
                FROM company_prices p
                JOIN companies c ON p.company_id = c.id
                WHERE c.code = %s
                  AND p.trade_date >= %s
            )
            SELECT 
                ROUND(MIN((low_price - running_max_high) / running_max_high * 100), 2) AS max_drawdown,
                ROUND(MIN((close_price - prev_close) / prev_close * 100), 2) AS worst_daily_drop,
                ROUND(AVG(CASE WHEN close_price < prev_close THEN (close_price - prev_close) / prev_close * 100 END), 2) AS avg_downside,
                ROUND(COUNT(CASE WHEN close_price < prev_close THEN 1 END) / COUNT(*) * 100, 2) AS down_prob
            FROM PeriodData
            WHERE prev_close IS NOT NULL;
            """
            
            logging.info(f"'{code}' 종목의 최근 {months}개월 리스크 지표를 계산합니다.")
            df = pd.read_sql(query, self.engine, params=(code, start_date_str))
            return df

        except Exception as e:
            logging.error(f"리스크 지표 계산 중 오류 발생: {e}")
            return pd.DataFrame()


def export_to_excel(dataframe, filename='golden_cross_stocks.xlsx'):
    """데이터프레임을 엑셀 파일로 저장합니다."""
    if dataframe.empty:
        logging.info("엑셀로 저장할 데이터가 없습니다.")
        return

    try:
        # reports 디렉터리가 없으면 생성
        if not os.path.exists('reports'):
            os.makedirs('reports')
        
        filepath = os.path.join('reports', filename)
        dataframe.to_excel(filepath, index=False)
        logging.info(f"검색 결과가 '{filepath}' 파일로 저장되었습니다.")
    except Exception as e:
        logging.error(f"엑셀 파일 저장 중 오류 발생: {e}")

if __name__ == "__main__":
    import logging
    from dotenv import load_dotenv
    load_dotenv()
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    with get_db_connection() as db_access:
        screener = Screener(db_access)
        screener.find_leading_stocks(top_n=50, min_marcap_억=500, require_ma_alignment=False)